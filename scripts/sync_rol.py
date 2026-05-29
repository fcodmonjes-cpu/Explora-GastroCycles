#!/usr/bin/env python3
"""
sync_rol.py — descarga el Excel de turnos desde SharePoint y actualiza Firebase.

Secrets requeridos (GitHub → Settings → Secrets → Actions):
  FIREBASE_KEY   →  contenido completo del JSON de la service account
  SHAREPOINT_URL →  link directo de descarga del Excel

Modo debug (no escribe en Firebase, imprime estructura):
  FIREBASE_KEY=$(cat firebase-key.json) SHAREPOINT_URL=<url> python scripts/sync_rol.py --debug
"""

import json, os, io, sys, re, datetime, unicodedata, requests
import openpyxl
from google.oauth2 import service_account
from google.auth.transport.requests import Request as GoogleAuthRequest

SHAREPOINT_URL = os.environ["SHAREPOINT_URL"]
DB_URL = "https://explora-cafe-orders-default-rtdb.firebaseio.com"
DEBUG  = "--debug" in sys.argv

# Alias: nombre en el Excel (tras limpiar sufijos) → nombre a mostrar en la app.
ALIASES = {
    "Francisco Monjes": "Bruno",
    # "Otro Nombre": "Apodo",
}

MES_VARIANTS = {
    1:  ["ENERO",      "ENE"],
    2:  ["FEBRERO",    "FEB"],
    3:  ["MARZO",      "MAR"],
    4:  ["ABRIL",      "ABR"],
    5:  ["MAYO",       "MAY"],
    6:  ["JUNIO",      "JUN"],
    7:  ["JULIO",      "JUL"],
    8:  ["AGOSTO",     "AGO"],
    9:  ["SEPTIEMBRE", "SEP", "SEPT"],
    10: ["OCTUBRE",    "OCT"],
    11: ["NOVIEMBRE",  "NOV"],
    12: ["DICIEMBRE",  "DIC"],
}

# Códigos de COMEDOR — los únicos que aparecen en staffing am/pm.
COMEDOR_AM  = {"CAM", "CAMC"}
COMEDOR_PM  = {"CPM", "CPMC"}
COMEDOR_DOB = {"CDOB"}

# Códigos que significan "no trabajando".
NOT_WORKING = {"L", "LA", "LIC", "V", "CU", "F", "x", ""}

# Roles que cuentan como "senior" en el comedor.
SENIOR_ROLES = {"GEO SENIOR", "SUP COMEDOR", "JEFE HOSP"}


# ── Firebase ──────────────────────────────────────────────────────────────────

def get_token():
    key_data = json.loads(os.environ["FIREBASE_KEY"])
    creds = service_account.Credentials.from_service_account_info(
        key_data,
        scopes=[
            "https://www.googleapis.com/auth/firebase.database",
            "https://www.googleapis.com/auth/userinfo.email",
        ],
    )
    creds.refresh(GoogleAuthRequest())
    return creds.token


def fb_put(token, path, data):
    r = requests.put(
        f"{DB_URL}/{path}.json",
        params={"access_token": token},
        json=data,
        timeout=20,
    )
    r.raise_for_status()


# ── Descarga ──────────────────────────────────────────────────────────────────

def download_excel():
    r = requests.get(
        SHAREPOINT_URL,
        headers={"User-Agent": "Mozilla/5.0"},
        allow_redirects=True,
        timeout=30,
    )
    r.raise_for_status()
    if r.content[:2] != b"PK":
        raise ValueError(
            f"La respuesta no es un xlsx ({len(r.content)} bytes). "
            "¿Expiró el link de SharePoint?"
        )
    return io.BytesIO(r.content)


# ── Utilidades de nombres ─────────────────────────────────────────────────────

def clean_name(raw):
    """Quita sufijos (TDP)/(VAL)/etc., aplica aliases y normaliza espacios."""
    if not raw:
        return None
    name = re.sub(r"\s*\([^)]*\)", "", str(raw)).strip()
    if not name:
        return None
    return ALIASES.get(name, name)


def norm_key(name):
    """Nombre sin tildes en minúscula — para deduplicar variantes del mismo nombre."""
    return unicodedata.normalize("NFD", name).encode("ascii", "ignore").decode().lower().strip()


# ── Clasificación de códigos ──────────────────────────────────────────────────

def comedor_franjas(code):
    """[], ['am'], ['pm'], o ['am','pm'] según el código de turno."""
    if not code:
        return []
    s = str(code).strip().upper()
    if s in COMEDOR_AM:  return ["am"]
    if s in COMEDOR_PM:  return ["pm"]
    if s in COMEDOR_DOB: return ["am", "pm"]
    return []


def apoyo_franjas(code):
    """Clasifica códigos informales de la sección Apoyos. Ignora bar/recepción."""
    if not code:
        return []
    s = str(code).strip().lower().replace(" ", "").replace("/", "")
    if s in {"bpm", "bam", "bpis", "rpm", "rsal", "rnoc",
             "bpmc", "bamc", "bpmapoyo", "bpisbpm"}:
        return []
    if s.startswith("cam") or s == "almbpm":
        return ["am"]
    if s in {"cpm", "cpmc", "cen", "cena"}:
        return ["pm"]
    if s == "dob":
        return ["am", "pm"]
    return []


def is_working(code):
    if not code:
        return False
    return str(code).strip().upper() not in NOT_WORKING


# ── Búsqueda de hoja ──────────────────────────────────────────────────────────

def find_sheet(wb, month: int):
    """
    Devuelve la hoja del mes. Prefiere el nombre base (MAY, JUNIO) sobre v2,
    porque la hoja base contiene la sección de Apoyos y los viajeros correctos.
    Cae en v2 solo si no existe la base.
    """
    for mes in MES_VARIANTS[month]:
        for candidate in (mes, f"{mes} v2"):
            if candidate in wb.sheetnames:
                print(f"[sync-rol] Usando hoja: {candidate}")
                return wb[candidate]
    raise ValueError(
        f"No se encontró hoja para mes {month}. "
        f"Hojas disponibles: {wb.sheetnames}"
    )


def get_day_cols(rows):
    """Detecta la fila DIA y retorna (dia_idx, day_cols {day_int: col_idx})."""
    dia_idx = next(
        (i for i, r in enumerate(rows) if len(r) > 9 and r[9] == "DIA"),
        None,
    )
    if dia_idx is None:
        raise ValueError("No se encontró fila 'DIA' en la hoja")
    day_cols = {
        int(v): ci
        for ci, v in enumerate(rows[dia_idx])
        if isinstance(v, (int, float)) and 1 <= v <= 31
    }
    return dia_idx, day_cols


# ── Parseo principal ──────────────────────────────────────────────────────────

def parse_excel(wb, month: int):
    """
    Retorna (staffing_dict, roster_dict).

    staffing: { "1": { viajeros, geo_senior_am, geo_senior_pm,
                        geos_am, geos_pm, apoyo_am, apoyo_pm }, ... }

    roster:   { "Nombre": { "1": "CAM", "5": "CPM", ... } }
    """
    ws   = find_sheet(wb, month)
    rows = list(ws.iter_rows(values_only=True))

    dia_idx, day_cols = get_day_cols(rows)

    # Viajeros (fila DIA + 2)
    occ_row = rows[dia_idx + 2]
    viajeros = {
        day: int(occ_row[col])
        for day, col in day_cols.items()
        if isinstance(occ_row[col], (int, float))
    }

    # ── Sección GEO / GEO SENIOR (col8 = rol explícito) ──────────────────────
    people = []   # lista de dicts: name, is_senior, comedor, codes
    for row in rows[dia_idx + 3:]:
        role_cell = row[8] if len(row) > 8 else None
        if role_cell not in (*SENIOR_ROLES, "GEO"):
            continue
        name = clean_name(row[9] if len(row) > 9 else None)
        if not name:
            continue

        comedor = {}
        codes   = {}
        for day, col in day_cols.items():
            code = row[col] if col < len(row) else None
            fr = comedor_franjas(code)
            if fr:
                comedor[day] = fr
            if is_working(code):
                codes[str(day)] = str(code).strip().upper()

        people.append({
            "name":      name,
            "norm":      norm_key(name),
            "is_senior": role_cell in SENIOR_ROLES,
            "comedor":   comedor,
            "codes":     codes,
        })
        if DEBUG:
            s = [f"{d}:{','.join(comedor.get(d,[])) or codes.get(str(d),'-')}"
                 for d in sorted(day_cols)[:7]]
            print(f"  {'SENIOR' if role_cell in SENIOR_ROLES else 'GEO':6s} {name}: {' '.join(s)}...")

    print(f"[sync-rol] {len(people)} personas GEO")

    # ── Sección Apoyos (col8 = None, col9 = nombre, días con códigos texto) ──
    apoyos = []
    for row in rows:
        if (row[8] if len(row) > 8 else None) is not None:
            continue
        name = clean_name(row[9] if len(row) > 9 else None)
        if not name:
            continue
        # Fila de persona si algún día tiene string (no número/None)
        day_vals = [row[col] if col < len(row) else None for col in day_cols.values()]
        if not any(isinstance(v, str) and v.strip() for v in day_vals):
            continue

        franjas_ap = {}
        codes_ap   = {}
        for day, col in day_cols.items():
            code = row[col] if col < len(row) else None
            fr = apoyo_franjas(code)
            if fr:
                franjas_ap[day] = fr
            if code and str(code).strip() and str(code).strip().lower() not in NOT_WORKING:
                codes_ap[str(day)] = str(code).strip().upper()

        if any(franjas_ap.values()):
            apoyos.append({
                "name":   name,
                "norm":   norm_key(name),
                "franjas": franjas_ap,
                "codes":   codes_ap,
            })
            if DEBUG:
                s = [f"{d}:{''.join(franjas_ap.get(d,[])) or '-'}"
                     for d in sorted(day_cols)[:7]]
                print(f"  APOYO  {name}: {' '.join(s)}...")

    print(f"[sync-rol] {len(apoyos)} personas Apoyos")

    # ── Construir staffing ────────────────────────────────────────────────────
    # Deduplicación por franja y por nombre normalizado:
    # se añade a "apoyo" solo si esa persona no aparece ya en geos_/geo_senior_
    # de esa misma franja (evita duplicados de GEOs que también están en apoyos).
    staffing = {}
    for day in sorted(day_cols):
        entry = {
            "viajeros":      viajeros.get(day, 0),
            "geo_senior_am": [],
            "geo_senior_pm": [],
            "geos_am":       [],
            "geos_pm":       [],
            "apoyo_am":      [],
            "apoyo_pm":      [],
        }
        seen_am = set()
        seen_pm = set()

        for p in people:
            for franja in p["comedor"].get(day, []):
                key = f"geo_senior_{franja}" if p["is_senior"] else f"geos_{franja}"
                entry[key].append(p["name"])
                (seen_am if franja == "am" else seen_pm).add(p["norm"])

        for ap in apoyos:
            for franja in ap["franjas"].get(day, []):
                seen = seen_am if franja == "am" else seen_pm
                if ap["norm"] not in seen:
                    entry[f"apoyo_{franja}"].append(ap["name"])
                    seen.add(ap["norm"])

        staffing[str(day)] = entry

    # ── Construir roster ──────────────────────────────────────────────────────
    # Formato que espera la app: { "Nombre": { "1": "CAM", "5": "CPM" } }
    # Los GEO usan sus códigos formales. Los apoyos puros usan sus códigos
    # de la sección de apoyos. Si coincide el nombre (normalizado), gana el GEO.
    geo_norms = {p["norm"]: p["name"] for p in people}
    roster = {p["name"]: p["codes"] for p in people if p["codes"]}
    for ap in apoyos:
        if ap["norm"] not in geo_norms and ap["codes"]:
            roster[ap["name"]] = ap["codes"]

    return staffing, roster


# ── Main ──────────────────────────────────────────────────────────────────────

def months_to_sync(today):
    """Mes actual + mes siguiente. Mantiene mayo visible mientras junio
    empieza a aparecer en cuanto el supervisor agregue la pestaña JUNIO."""
    cur = (today.year, today.month)
    nxt = (today.year + 1, 1) if today.month == 12 else (today.year, today.month + 1)
    return [cur, nxt]


def main():
    today = datetime.date.today()

    print(f"[sync-rol] {today} — descargando Excel...")
    xlsx = download_excel()
    print(f"[sync-rol] {xlsx.getbuffer().nbytes:,} bytes recibidos")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    parsed = {}  # month_str → (staffing, roster)
    for (y, m) in months_to_sync(today):
        month_str = f"{y:04d}-{m:02d}"
        try:
            staffing, roster = parse_excel(wb, m)
        except ValueError as e:
            print(f"[sync-rol] {month_str} — saltado: {e}")
            continue
        print(f"[sync-rol] {month_str}: {len(staffing)} dias · {len(roster)} personas en roster")
        parsed[month_str] = (staffing, roster)

    if not parsed:
        raise SystemExit("[sync-rol] No se pudo parsear ningún mes — abortando.")

    if DEBUG:
        import pprint
        cur_str = today.strftime("%Y-%m")
        if cur_str in parsed:
            staffing, roster = parsed[cur_str]
            print(f"\n--- staffing {cur_str} dia {today.day} ---")
            pprint.pprint(staffing.get(str(today.day)))
            print("\n--- roster Bruno ---")
            pprint.pprint(roster.get("Bruno"))
            print("\n--- roster Bryan Rincon ---")
            pprint.pprint(roster.get("Bryan Rincon"))
        print(f"\n[sync-rol] Meses parseados: {sorted(parsed.keys())}")
        print("[sync-rol] Modo debug — Firebase no modificado.")
        return

    print("[sync-rol] Autenticando con Firebase...")
    token = get_token()

    for month_str, (staffing, roster) in parsed.items():
        fb_put(token, f"staffing/{month_str}", staffing)
        print(f"[sync-rol] OK /staffing/{month_str}")
        fb_put(token, f"roster/{month_str}", roster)
        print(f"[sync-rol] OK /roster/{month_str}")

    now_utc = datetime.datetime.utcnow()
    fb_put(token, "meta/last_update", {
        "timestamp_iso":   now_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "timestamp_local": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "timestamp_ms":    int(now_utc.timestamp() * 1000),
        "mes":             ",".join(sorted(parsed.keys())),
        "tipo":            "sync automatico",
    })
    print(f"[sync-rol] OK /meta/last_update")
    print("[sync-rol] Listo.")


if __name__ == "__main__":
    main()
